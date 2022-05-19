# ==============================================================================
#
#          FILE:  L999999999_0.py
#
#         USAGE:  It's a python library
#                    from L999999999_0 import (
#                        hxltm_carricato,
#                        qhxl_hxlhashtag_2_bcp47,
#                        # (...)
#                    )
#
#   DESCRIPTION:  Common library for 999999999_*.py cli scripts
#
#       OPTIONS:  ---
#
#  REQUIREMENTS:  - python3
#                   - pip install openpyxl
#          BUGS:  ---
#         NOTES:  ---
#       AUTHORS:  Emerson Rocha <rocha[at]ieee.org>
# COLLABORATORS:  ---
#       COMPANY:  EticaAI
#       LICENSE:  Public Domain dedication or Zero-Clause BSD
#                 SPDX-License-Identifier: Unlicense OR 0BSD
#       VERSION:  v1.0.0
#       CREATED:  2022-05-18 19:52 UTC based on 999999999_10263485.py
#      REVISION:  ---
# ==============================================================================
"""Common library for 999999999_*.py cli scripts
"""

import csv
import re
import sys
import datetime
from typing import Type

from openpyxl import (
    load_workbook
)

# pylint --disable=W0511,C0103,C0116 ./999999999/0/L999999999_0.py


def csv_imprimendo(
        data: list, caput: list = None, delimiter: str = ',',
        archivum_trivio: str = None):
    if archivum_trivio:
        raise NotImplementedError('{0}'.format(archivum_trivio))
    # imprimendō, v, s, dativus, https://en.wiktionary.org/wiki/impressus#Latin

    _writer = csv.writer(sys.stdout, delimiter=delimiter)
    if caput:
        _writer.writerow(caput)
    _writer.writerows(data)


def hxltm_carricato(
    archivum_trivio: str = None,
    est_stdin: bool = False
) -> list:
    """hxltm_carricato load entire raw CSV file to memory.

    Note: this helper is not as efficent as read line by line. But some
    operations already require such task.

    Trivia:
    - carricātō, n, s, dativus, https://en.wiktionary.org/wiki/carricatus#Latin
      - verbum: https://en.wiktionary.org/wiki/carricatus#Latin

    Args:
        archivum_trivio (str, optional): Path to file. Defaults to None.
        est_stdin (bool, optional): Is the file stdin?. Defaults to False.

    Returns:
        list: list of [caput, data], where data is array of lines
    """
    caput = []
    _data = []

    if est_stdin:
        for linea in sys.stdin:
            if len(caput) == 0:
                # caput = linea
                # _reader_caput = csv.reader(linea)
                _gambi = [linea, linea]
                _reader_caput = csv.reader(_gambi)
                caput = next(_reader_caput)
            else:
                _data.append(linea)
        _reader = csv.reader(_data)
        return caput, list(_reader)
    # else:
    #     fons = archivum_trivio

    with open(archivum_trivio, 'r') as _fons:
        _csv_reader = csv.reader(_fons)
        for linea in _csv_reader:
            if len(caput) == 0:
                # caput = linea
                # _reader_caput = csv.reader(linea)
                _gambi = [linea, linea]
                _reader_caput = csv.reader(_gambi)
                caput = next(_reader_caput)
            else:
                _data.append(linea)
            # print(row)

    # for line in fons:
    #     print(line)
        # json_fonti_texto += line

    _reader = csv.reader(_data)
    return caput, list(_reader)


def qhxl_hxlhashtag_2_bcp47(hxlhashtag: str) -> str:
    """qhxl_hxlhashtag_2_bcp47

    (try) to convert full HXL hashtag to BCP47

    Args:
        hxlatt (str):

    Returns:
        str:
    """
    # needs simplification
    if not hxlhashtag:
        return None
    if hxlhashtag.find('i_') == -1 or hxlhashtag.find('is_') == -1:
        return None
    hxlhashtag_parts = hxlhashtag.split('+')
    # langattrs = []
    _bcp_lang = ''
    _bcp_stript = ''
    _bcp_extension = []
    for item in hxlhashtag_parts:
        if item.startswith('i_'):
            _bcp_lang = item.replace('i_', '')
        if item.startswith('is_'):
            _bcp_stript = item.replace('is_', '')
        if item.startswith('ix_'):
            _bcp_extension.append(item.replace('ix_', ''))
        # if not item.startswith(('i_', 'is_', 'ix_')):
        #     continue
        # langattrs.append(item)

    if not _bcp_lang or not _bcp_stript:
        return False

    bcp47_simplici = "{0}-{1}".format(
        _bcp_lang.lower(), _bcp_stript.capitalize())
    if len(_bcp_extension) > 0:
        _bcp_extension = sorted(_bcp_extension)
        bcp47_simplici = "{0}-x-{1}".format(
            bcp47_simplici,
            '-'.join(_bcp_extension)
        )

    return bcp47_simplici


class XLSXSimplici:
    """Read-only wrapper for XLSX files

    - XLSX http://officeopenxml.com/anatomyofOOXML-xlsx.php
    - simplicī, m/f/n, s, dativus, https://en.wiktionary.org/wiki/simplex#Latin
    """

    archivum_trivio: str = ''
    workbook: None
    active: str = None
    active_col_start: int = 0
    active_col_end: int = 1
    active_col_ignore: list = []  # Example: headers without text
    active_row_start: int = 0
    active_row_end: int = 1
    _formatum: str = 'csv'

    def __init__(self, archivum_trivio: str) -> None:
        """__init__

        Args:
            archivum_trivio (str):
        """
        # from openpyxl import load_workbook
        self.archivum_trivio = archivum_trivio
        self.workbook = load_workbook(
            archivum_trivio, data_only=True, read_only=True)
        # self.workbook = load_workbook(
        #     archivum_trivio, data_only=True)

        self.workbook.iso_dates = True
        self.sheet_active = None
        # pass

    def de(self, worksheet_reference: str = None) -> bool:
        if isinstance(worksheet_reference, str) and \
                len(worksheet_reference) == 1:
            meta = self.meta()
            if worksheet_reference in meta['sheet_cod_ab']:
                self.active = meta['sheet_cod_ab'][worksheet_reference]
                return True
        if worksheet_reference in self.workbook:
            self.active = worksheet_reference
            return True
        return False

    def meta(self) -> dict:
        """meta

        Returns:
            dict: _description_
        """
        resultatum = {
            # '_': self.workbook,
            'sheetnames': self.workbook.sheetnames,
            'sheet': {},
            'sheet_active': {},
            'sheet_cod_ab': {}
        }
        for item in self.workbook.sheetnames:
            resultatum['sheet'][item] = {
                # '__': self.workbook[item],
                # 'max_col': self.workbook[item].max_col
                'max_column': self.workbook[item].max_column,
                'max_row': self.workbook[item].max_row
            }
            _item_num = re.sub('[^0-9]', '', item)
            if len(_item_num) == 1:
                #_likely_ab = _item_num
                resultatum['sheet_cod_ab'][_item_num] = item
        # resultatum = wb2
        # print(wb2.keys())
        # workbook.close()
        return resultatum

    def finis(self) -> None:
        """finis Close XLSX file immediately to save memory
        """
        # https://en.wiktionary.org/wiki/finis#Latin
        self.workbook.close()

    def imprimere(self, formatum: str = None) -> list:
        """imprimere /print/@eng-Latn

        Trivia:
        - imprimere, v, s, (), https://en.wiktionary.org/wiki/imprimo#Latin
        - fōrmātum, s, n, nominativus, https://en.wiktionary.org/wiki/formatus

        Args:
            formatum (str, optional): output format. Defaults to 'csv'.

        Returns:
            [list]: data, caput
        """
        caput = []
        data = []
        if formatum:
            self._formatum = formatum

        fons = self.workbook[self.active]

        for row_index, row in enumerate(fons.rows):
            if row_index >= self.active_row_start and \
                    row_index <= self.active_row_end:
                linea = []
                for col_index in range(
                        self.active_col_start, self.active_col_end -1):
                    textum = row[col_index].value

                    if textum:
                        if isinstance(textum, datetime.datetime):
                            textum = str(textum).replace(' 00:00:00', '')
                        linea.append(textum)
                    else:
                        linea.append('')

                if len(caput) == 0:
                    caput = linea
                else:
                    data.append(linea)
            # a_dict[row[0].value+','+row[1].value].append(str(row[2].value))

        return data, caput 

    def praeparatio(self):
        """praeparātiō

        Trivia:
        - praeparātiō, s, f, Nom., https://en.wiktionary.org/wiki/praeparatio
        """

        self.active_col_end = self.workbook[self.active].max_column
        self.active_row_end = self.workbook[self.active].max_row

        fons = self.workbook[self.active]

        for row_index, row in enumerate(fons.rows):
            # TODO: discover where it starts
            pass

        # for index_cols in fons.rows

        # @TODO: calculate if HXL, if have more columns than ideal, etc

        return True
